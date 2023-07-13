# скорее всего это потом станет прогой
import requests
from bs4 import BeautifulSoup as bs
import urllib.request
import openpyxl
import telebot
import os
from DAO import *
url='https://ba.hse.ru/base2023'
progs=set()
DB=DataBase()
DB.create_table()
def get_statistic():
    response = requests.get(url)
    soup = bs(response.text, 'lxml')

    blocks = soup.find_all('div', class_='builder-section foldable_block')

    msk_info = blocks[0]
    info_all_link = msk_info.find_all('a')[0].get('href')
    urllib.request.urlretrieve(info_all_link, "statistic.xlsx")
def find_program_list(need_find="Прикладная математика и информатика"):
    response = requests.get(url)
    soup = bs(response.text, 'lxml')

    blocks = soup.find_all('div', class_='builder-section foldable_block')

    msk_info = blocks[0]
    program_info = []
    for link in msk_info.find_all('tr'):
        program_info.append(link)


    for i in range(len(program_info)):
        name=program_info[i].find_all('td')[0].text
        if (name == need_find):
            link=program_info[i].find_all('td')[1].find_all('a')[0].get('href')
            urllib.request.urlretrieve(link, "list_abi.xlsx")
            break
#начинаем обработку xlsx файлов
def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]
def get_info_statistic(path, target="Прикладная математика и информатика"):
    wb_obj = openpyxl.load_workbook(path)
    ws=wb_obj.active
    for merge in list(ws.merged_cells):
        ws.unmerge_cells(range_string=str(merge))
    res=list(iter_rows(ws))
    #print(res)
    titles=res[6]
    res_mes=""
    for row in res:
        if (row[1]==target):
            #print(res[3][0])
            res_mes=res_mes+res[3][0]+'\n'
            #print(res[6][10])
            res_mes = res_mes + res[6][10] + '\n'
            #print(row[10])
            res_mes = res_mes + row[10] + '\n'
            #print(res[6][14][8:])
            res_mes = res_mes + res[6][14][8:] + '\n'
            #print(row[14])
            res_mes = res_mes + row[14] + '\n'
            return res_mes
def get_place(path, snils):
    wb_obj = openpyxl.load_workbook(path)
    ws = wb_obj.active
    res_mes=""
    for merge in list(ws.merged_cells):
        ws.unmerge_cells(range_string=str(merge))
    res = list(iter_rows(ws))
    for row in res:
        if (row[2]==snils):
            #print(res[9][2])
            res_mes=res_mes+res[9][2]+'\n'
            res_mes=res_mes+"Место в списке абитуриентов  {0} :  {1}".format(res[1][2], row[0])+'\n'
            return res_mes
    return "Вас нет в списке абитуриентов данной программы"
API_KEY=""
with open('API_KEY','r') as f:
    API_KEY=f.readline()
API_KEY=API_KEY[:-1]
bot=telebot.TeleBot(API_KEY)
@bot.message_handler(commands=['start'])
def hello(message):
    bot.send_message(message.chat.id,"Запуск произведен для пользователя {0}".format(message.chat.id))
    if not(DB.check_exist(message.chat.id)):
        bot.send_message(message.chat.id, "Отправьте свой номер СНИЛС в формате XXX-XXX-XXX XX")
        bot.register_next_step_handler(message, add_snils)
def add_snils(message):
    #add_users(message.chat.id,message.text)
    snils=message.text
    bot.send_message(message.chat.id, "Введите интересующую программу")
    bot.register_next_step_handler(message, add_program, snils)
def check_program(program):
    if (len(progs) == 0):
        with open("prog_list.txt","r") as f:
            data=f.read()
            for line in data.splitlines():
                #s=s[:-1]
                progs.add(line)
    if (program in progs):
        return True
    else:
        return False
def add_program(message, snils):
    program=message.text
    if not(check_program(program)):
        bot.send_message(message.chat.id,"Такой программы не существует")
        bot.send_message(message.chat.id, "Введите интересующую программу")
        bot.register_next_step_handler(message,add_program, snils)
        return
    DB.add_user(message.chat.id, snils, program)
    print("Done adding")
@bot.message_handler(commands=['place'])
def get_place_pos(message):
    print("place_request_open")
    user=DB.get_from_table(message.chat.id)
    #snils=get_from_table(message.chat.id)[2]
    #program=
    find_program_list(user[3])
    res = get_place("list_abi.xlsx", user[2])
    bot.send_message(message.chat.id, res)
    os.remove("list_abi.xlsx")
    print("place_request_close")
@bot.message_handler(commands=['statistic'])
def get_statistic_b(message):
    print("statistic_request_open")
    get_statistic()
    user=DB.get_from_table(message.chat.id)
    res = get_info_statistic("statistic.xlsx",user[3])
    bot.send_message(message.chat.id, res)
    os.remove("statistic.xlsx")
    print("statistic_request_close")
@bot.message_handler(commands=['set_snils'])
def getu_snils(message):
    bot.send_message(message.chat.id, "Отправьте свой номер СНИЛС в формате XXX-XXX-XXX XX")
    bot.register_next_step_handler(message,ru_snils)
def ru_snils(message):
    DB.set_snils(message.chat.id,message.text)
@bot.message_handler(commands=['set_program'])
def getu_program(message):
    bot.send_message(message.chat.id, "Введите интересующую программу")
    bot.register_next_step_handler(message,ru_program)
def ru_program(message):
    program = message.text
    if not (check_program(program)):
        bot.send_message(message.chat.id, "Такой программы не существует")
        bot.send_message(message.chat.id, "Введите интересующую программу")
        bot.register_next_step_handler(message, ru_program)
        return
    DB.set_program(message.chat.id,program)
    print("Done adding")
bot.polling()